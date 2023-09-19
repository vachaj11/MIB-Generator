"""This module allows for generation of excel table mirroring the mib databases with a specific formatting.

The main generation processes happen in the module :obj:`gen_xls` from which other modules are called each of which (of those whose
names start with "gen_") adds one worksheet to the workbook usually containing information from one or two MIB tables. There are
also a few specific methods here (which don't start with "gen_") which only do some specific modification to the sheets unrelated
to their content.

Attributes:
    h1 (openpyxl.styles.NamedStyle): One of 8 ``openpyxl`` styles included as global attributes here which can be applied to cells.

Note:
    Everything here is very much not finished.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle

import os

import mib_generator.data.longdata as longdata

h1 = NamedStyle(name="h1")
h1.font = Font(bold=True)
h1.alignment = Alignment(horizontal="center", vertical="center")
h1.fill = PatternFill(fill_type="solid", fgColor="99CCFF")
h2 = NamedStyle(name="h2")
h2.font = Font(bold=True)
h2.alignment = Alignment(horizontal="center", vertical="center")
h2.fill = PatternFill(fill_type="solid", fgColor="CC99FF")
h3 = NamedStyle(name="h3")
h3.font = Font(bold=True)
h3.alignment = Alignment(horizontal="center", vertical="center")
h3.fill = PatternFill(fill_type="solid", fgColor="FFCC00")
h4 = NamedStyle(name="h4")
h4.font = Font(bold=True)
h4.alignment = Alignment(horizontal="center", vertical="center")
h4.fill = PatternFill(fill_type="solid", fgColor="FF8080")

c1 = NamedStyle(name="c1")
c1.font = Font(bold=False)
c1.alignment = Alignment(horizontal="left", vertical="center")
c1.fill = PatternFill(fill_type="solid", fgColor="99CCFF")
c2 = NamedStyle(name="c2")
c2.font = Font(bold=False)
c2.alignment = Alignment(horizontal="left", vertical="center")
c2.fill = PatternFill(fill_type="solid", fgColor="CC99FF")
c3 = NamedStyle(name="c3")
c3.font = Font(bold=False)
c3.alignment = Alignment(horizontal="left", vertical="center")
c3.fill = PatternFill(fill_type="solid", fgColor="FFCC00")
c4 = NamedStyle(name="c4")
c4.font = Font(bold=False)
c4.alignment = Alignment(horizontal="left", vertical="center")
c4.fill = PatternFill(fill_type="solid", fgColor="FF8080")


def gen_polyc(book, calib):
    ws = book.create_sheet("TM PolyCal")
    head_gen(ws, longdata.xls_heads["TM PolyCal"])
    for i in range(len(calib)):
        ws.cell(i + 3, 1, "PolyCal").style = c1
        ws.cell(i + 3, 2, calib[i].mcf["MCF_IDENT"]).style = c1
        ws.cell(i + 3, 3, calib[i].mcf["MCF_DESCR"]).style = c1
        ws.cell(i + 3, 4, calib[i].mcf["MCF_POL1"]).style = c1
        ws.cell(i + 3, 5, calib[i].mcf["MCF_POL2"]).style = c1
        ws.cell(i + 3, 6, calib[i].mcf["MCF_POL3"]).style = c1
        ws.cell(i + 3, 7, calib[i].mcf["MCF_POL4"]).style = c1
        ws.cell(i + 3, 8, calib[i].mcf["MCF_POL5"]).style = c1


def gen_logc(book, calib):
    ws = book.create_sheet("TM LogCal")
    head_gen(ws, longdata.xls_heads["TM LogCal"])
    for i in range(len(calib)):
        ws.cell(i + 3, 1, "LogCal").style = c1
        ws.cell(i + 3, 2, calib[i].lgf["LGF_IDENT"]).style = c1
        ws.cell(i + 3, 3, calib[i].lgf["LGF_DESCR"]).style = c1
        ws.cell(i + 3, 4, calib[i].lgf["LGF_POL1"]).style = c1
        ws.cell(i + 3, 5, calib[i].lgf["LGF_POL2"]).style = c1
        ws.cell(i + 3, 6, calib[i].lgf["LGF_POL3"]).style = c1
        ws.cell(i + 3, 7, calib[i].lgf["LGF_POL4"]).style = c1
        ws.cell(i + 3, 8, calib[i].lgf["LGF_POL5"]).style = c1


def gen_txtc(book, calib):
    ws = book.create_sheet("TM TxtCal")
    head_gen(ws, longdata.xls_heads["TM TxtCal"])
    ind = 4
    for i in calib:
        ws.cell(ind, 1, "TxtCal").style = c1
        ws.cell(ind, 2, i.txf["TXF_NUMBR"]).style = c1
        ws.cell(ind, 3, i.txf["TXF_DESCR"]).style = c1
        ws.cell(ind, 4, i.txf["TXF_NALIAS"]).style = c1
        ws.cell(ind, 5, i.txf["TXF_RAWFMT"]).style = c1
        ind += 1
        for l in i.txp:
            ws.cell(ind, 1, "TxtCal Ranges").style = c2
            ws.cell(ind, 3, l["TXP_FROM"]).style = c2
            ws.cell(ind, 4, l["TXP_TO"]).style = c2
            ws.cell(ind, 5, l["TXP_ALTXT"]).style = c2
            ind += 1
        ind += 1


def gen_numc(book, calib):
    ws = book.create_sheet("TM NumCal")
    head_gen(ws, longdata.xls_heads["TM NumCal"])
    ind = 4
    for i in calib:
        ws.cell(ind, 1, "NumCal").style = c1
        ws.cell(ind, 2, i.caf["CAF_NUMBR"]).style = c1
        ws.cell(ind, 3, i.caf["CAF_DESCR"]).style = c1
        ws.cell(ind, 4, i.caf["CAF_NCURVE"]).style = c1
        ws.cell(ind, 5, i.caf["CAF_RAWFMT"]).style = c1
        ws.cell(ind, 6, i.caf["CAF_RADIX"]).style = c1
        ws.cell(ind, 7, i.caf["CAF_ENGFMT"]).style = c1
        ws.cell(ind, 8, i.caf["CAF_UNIT"]).style = c1
        ws.cell(ind, 9, i.caf["CAF_INTER"]).style = c1
        ind += 1
        for l in i.cap:
            ws.cell(ind, 1, "NumCal Points").style = c2
            ws.cell(ind, 3, l["CAP_XVALS"]).style = c2
            ws.cell(ind, 4, l["CAP_YVALS"]).style = c2
            ind += 1
        ind += 1


def gen_tdec(book, calib):
    ws = book.create_sheet("TC TxtCal")
    head_gen(ws, longdata.xls_heads["TC TxtCal"])
    ind = 4
    for i in calib:
        ws.cell(ind, 1, "TC TxtCal").style = c1
        ws.cell(ind, 2, i.paf["PAF_NUMBR"]).style = c1
        ws.cell(ind, 3, i.paf["PAF_DESCR"]).style = c1
        ws.cell(ind, 4, i.paf["PAF_RAWFMT"]).style = c1
        ws.cell(ind, 5, i.paf["PAF_NALIAS"]).style = c1
        ind += 1
        for l in i.pas:
            ws.cell(ind, 1, "TC TxtCal Alias").style = c2
            ws.cell(ind, 3, l["PAS_ALVAL"]).style = c2
            ws.cell(ind, 4, l["PAS_ALTXT"]).style = c2
            ind += 1
        ind += 1


def gen_verif(book, verif):
    ws = book.create_sheet("TC Verif Criteria")
    head_gen(ws, longdata.xls_heads["TC Verif Criteria"])
    ind = 4
    for i in verif:
        ws.cell(ind, 1, "TC Verif Stage").style = c1
        ws.cell(ind, 2, i.cvs["CVS_ID"]).style = c1
        ws.cell(ind, 3, i.cvs["CVS_TYPE"]).style = c1
        ws.cell(ind, 4, i.cvs["CVS_SOURCE"]).style = c1
        ws.cell(ind, 5, i.cvs["CVS_START"]).style = c1
        ws.cell(ind, 5, i.cvs["CVS_INTERVAL"]).style = c1
        ind += 1


def gen_pidt(book, packets):
    ws = book.create_sheet("TM Packet")
    head_gen(ws, longdata.xls_heads["TM Packet"])
    ind = 5
    for i in packets:
        ws.cell(ind, 1, "TM Packet").style = c1
        ws.cell(ind, 2, i.tpcf["TPCF_NAME"]).style = c1
        ws.cell(ind, 3, i.pid["PID_SPID"]).style = c1
        ws.cell(ind, 5, i.pid["PID_APID"]).style = c1
        ws.cell(ind, 6, i.pid["PID_PI1_VAL"]).style = c1
        # ws.cell(ind, 7, i.pid["PID_PI2_VAL"]).style = c1
        ws.cell(ind, 11, i.pid["PID_DESCR"]).style = c1
        ind += 1
        for l in range(len(i.entries)):
            if i.pcf[l]["PCF_NAME"] in [k["VPD_NAME"] for k in i.vpd]:
                vpd = next((x for x in i.vpd if x["VPD_NAME"] == i.pcf[l]["PCF_NAME"]))
                ws.cell(ind, 1, "TM Packet Variable Element").style = c3
                ws.cell(ind, 4, vpd["VPD_NAME"]).style = c3
                ws.cell(ind, 5, vpd["VPD_POS"]).style = c3
                ws.cell(ind, 6, vpd["VPD_GRPSIZE"]).style = c3
                ws.cell(ind, 7, vpd["VPD_FIXREP"]).style = c3
                # ws.cell(ind, 10, vpd["VPD_DIDESC"]).style = c3
                ws.cell(ind, 11, vpd["VPD_WIDTH"]).style = c3
            else:
                ws.cell(ind, 1, "TM Packet Fixed Element").style = c2
                ws.cell(ind, 3, i.pcf[l]["PCF_NAME"]).style = c2
                ws.cell(ind, 4, i.plf[l]["PLF_OFFBY"]).style = c2
                ws.cell(ind, 5, i.plf[l]["PLF_OFFBI"]).style = c2
            ind += 1
        ind += 1

def gen_ccft(book, packets):
    ws = book.create_sheet("TC List")
    head_gen(ws, longdata.xls_heads["TC List"])
    ind = 6
    
    for i in packets:
        ws.cell(ind, 1, "TC").style = c1
        ws.cell(ind, 2, i.ccf["CCF_CNAME"]).style = c1
        ws.cell(ind, 3, i.ccf["CCF_DESCR"]).style = c1
        ws.cell(ind, 4, i.ccf["CCF_DESCR2"]).style = c1
        #ws.cell(ind, 5, i.ccf["CCF_CTYPE"]).style = c1
        ws.cell(ind, 6, i.ccf["CCF_TYPE"]).style = c1
        ws.cell(ind, 7, i.ccf["CCF_STYPE"]).style = c1
        ws.cell(ind, 8, i.ccf["CCF_APID"]).style = c1
        ws.cell(ind, 9, i.ccf["CCF_NPARS"]).style = c1
        #ws.cell(ind, 10, i.ccf["CCF_CRITICAL"]).style = c1
        #ws.cell(ind, 11, i.ccf["CCF_SUBSYS"]).style = c1
        ws.cell(ind, 12, i.ccf["CCF_PKTID"]).style = c1
        #ws.cell(ind, 13, i.ccf["CCF_PLAN"]).style = c1
        #ws.cell(ind, 14, i.ccf["CCF_EXEC"]).style = c1
        #ws.cell(ind, 15, i.ccf["CCF_ILSCOPE"]).style = c1
        #ws.cell(ind, 16, i.ccf["CCF_ILSTAGE"]).style = c1
        #ws.cell(ind, 17, i.ccf["CCF_HIPRI"]).style = c1
        #ws.cell(ind, 18, i.ccf["CCF_MAPID"]).style = c1
        #ws.cell(ind, 19, i.ccf["CCF_DEFSET"]).style = c1
        #ws.cell(ind, 20, i.ccf["CCF_RAPID"]).style = c1
        #ws.cell(ind, 21, i.ccf["CCF_ACK"]).style = c1
        #ws.cell(ind, 22, i.ccf["CCF_SUBSCHEDID"]).style = c1
        ws.cell(ind, 23, i.ccf["CCF_DESCR2"]).style = c1
        ind += 1
        for l in i.cdf:
            ws.cell(ind, 1, "TC Element").style = c2
            ws.cell(ind, 3, l["CDF_CNAME"]).style = c2
            ws.cell(ind, 4, l["CDF_VALUE"]).style = c2
            ws.cell(ind, 5, l["CDF_ELTYPE"]).style = c2
            ws.cell(ind, 6, l["CDF_DESCR"]).style = c2
            ws.cell(ind, 7, l["CDF_ELLEN"]).style = c2
            ws.cell(ind, 8, l["CDF_BIT"]).style = c2
            ws.cell(ind, 9, l["CDF_GRPSIZE"]).style = c2
            #ws.cell(ind, 11, l["CDF_TMID"]).style = c2
            ind +=1
        for l in i.cvp:
            ws.cell(ind, 1, "TC Verif Profile").style = c4
            ws.cell(ind, 3, l["CVP_CVSID"]).style = c4
            ind += 1
        ind += 1
            
        

def gen_xls(
    Tm_packets=None,
    Tc_packets=None,
    calibrations=None,
    decalibrations=None,
    verifications=None,
    Tc_head=None,
):
    """The main function generating the excel table.
    
    It first calls various other "gen_" functions, each of which adds one worksheet to the workbook based on
    passed packets/calibrations/etc..., then a blank sheet that was initially created with the workbook is
    deleted and the method checks whether all expected worksheets have already been created (based on information
    in :obj:`mib_generator.data.longdata.xls_heads`). If not, then empty sheets containing only the table headers
    are created. Finally the width of columns in all sheets is adjusted and the workbook is returned.
    
    Args:
        Tm_packets (list): List of TM-packets (their Python representations), each represented by an object of type
            :obj:`mib_generator.construction.TM_packet.TM_packet`.
        Tc_packets (list): List of TC-commands (their Python representations), each represented by an object of type
            :obj:`mib_generator.construction.TC_packet.TC_packet`.
        calibrations (dict): Dictionary holding lists various calibrations (their Python representations), each calibration
            being an object of some child-class of :obj:`mib_generator.construction.calib.calib`.
        decalibrations (list): List of textual decalibrations (their Python representations), each represented by an object of
            type :obj:`mib_generator.construction.calib.decalib`.
        verifications (list): List of TC-command verifications (their Python representations), each represented by an object of
            type :obj:`mib_generator.construction.calib.verification`.
        Tc_head (construction.TC_packet.TC_header): A header included in all TC-commands.
        
    Returns:
        openpyxl.Workbook: An excel workbook containing all the generated sheets/tables.
    """
    wb = Workbook()
    gen_polyc(wb, calibrations["mcfs"])
    gen_logc(wb, calibrations["lgfs"])
    gen_txtc(wb, calibrations["txfs"])
    gen_numc(wb, calibrations["cafs"])
    gen_tdec(wb, decalibrations)
    gen_verif(wb, verifications)
    gen_pidt(wb, Tm_packets)
    gen_ccft(wb, Tc_packets)
    
    wb.remove(wb["Sheet"])
    
    for i in longdata.xls_heads:
        if i not in wb.sheetnames:
            ws = wb.create_sheet(i)
            head_gen(ws, longdata.xls_heads[i])
    
    for i in wb:
        adjust_width(i)
    return wb


def adjust_width(ws):
    for column in ws.column_dimensions:
        column.width = 3
    for row in ws.rows:
        for cell in row:
            if cell.value:
                ws.column_dimensions[cell.column_letter].width = max(
                    (
                        ws.column_dimensions[cell.column_letter].width,
                        len(str(cell.value)) * 1.2,
                    )
                )


def head_gen(sheet, heads):
    for i in range(len(heads)):
        for l in heads[i]:
            if l:
                sheet.cell(i + 1, heads[i].index(l) + 1, l).style = globals()[
                    "h" + str(i + 1)
                ]
